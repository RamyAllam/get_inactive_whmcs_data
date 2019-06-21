import pymysql
import xlsxwriter
from openpyxl import load_workbook
from vars import *
from queries import inactive_clients_query
from VerifyEmailAddress import verify_email

# Open XLSX file and adding sheets
workbook = xlsxwriter.Workbook(export_file)

# Formatting cells
cells_titles_format = workbook.add_format({'bold': True})
cells_titles_format.set_bg_color("#81BEF7")

worksheet_results_details = workbook.add_worksheet("Results")

# Set width for columns
worksheet_results_details.set_column(user_id_col, user_id_col, 10)
worksheet_results_details.set_column(firstname_col, firstname_col, 20)
worksheet_results_details.set_column(lastname_col, lastname_col, 20)
worksheet_results_details.set_column(email_col, email_col, 40)
worksheet_results_details.set_column(address1_col, address1_col, 40)
worksheet_results_details.set_column(city_col, city_col, 20)
worksheet_results_details.set_column(state_col, state_col, 20)
worksheet_results_details.set_column(country_col, country_col, 10)
worksheet_results_details.set_column(whmcs_url_col, whmcs_url_col, 30)


# Format to write ( row, column, content, format )
worksheet_results_details.write(0, user_id_col, "User ID", cells_titles_format)
worksheet_results_details.write(0, firstname_col, "First Name", cells_titles_format)
worksheet_results_details.write(0, lastname_col, "Last Name", cells_titles_format)
worksheet_results_details.write(0, email_col, "Email", cells_titles_format)
worksheet_results_details.write(0, address1_col, "Address", cells_titles_format)
worksheet_results_details.write(0, city_col, "City", cells_titles_format)
worksheet_results_details.write(0, state_col, "State", cells_titles_format)
worksheet_results_details.write(0, country_col, "Country", cells_titles_format)
worksheet_results_details.write(0, whmcs_url_col, "WHMCS Profile", cells_titles_format)


# Open database connection
db = pymysql.connect(db_host, db_user, db_pass, db_name )

# prepare a cursor object using cursor() method
cursor = db.cursor()

email_and_row_dict = dict()
try:
    # execute SQL query using execute() method.
    cursor.execute(inactive_clients_query)

    # Fetch all the rows in a list of lists.
    results = cursor.fetchall()

    for row in results:
        row_count += 1
        id = row[0]
        firstname = row[1]
        lastname = row[2]
        email = row[3]
        address1 = row[4]
        city = row[5]
        state = row[6]
        country = row[7]
        worksheet_results_details.write(row_count, user_id_col, id)
        worksheet_results_details.write(row_count, firstname_col, firstname)
        worksheet_results_details.write(row_count, lastname_col, lastname)
        worksheet_results_details.write(row_count, email_col, email)
        worksheet_results_details.write(row_count, address1_col, address1)
        worksheet_results_details.write(row_count, city_col, city)
        worksheet_results_details.write(row_count, state_col, state)
        worksheet_results_details.write(row_count, country_col, country)
        worksheet_results_details.write(
            row_count, whmcs_url_col, "{}={}".format(whmcs_user_url, id)
        )

        if wanna_verify_email:
            worksheet_results_details.set_column(email_valid_col, email_valid_col, 10)
            worksheet_results_details.write(0, email_valid_col, "Email Valid", cells_titles_format)

            email_and_row_dict[row_count] = email

except Exception as e:
    print("Exception: {}".format(e))

# disconnect from server
db.close()

workbook.close()

if wanna_verify_email:
    wb = load_workbook(filename=export_file)
    ws = wb.get_active_sheet()
    for row, email in email_and_row_dict.items():
        email_verification_results = verify_email(from_address, email)
        ws.cell(row=row+1, column=email_valid_col+1).value = str(email_verification_results)
        print(row, email)
    wb.save(export_file)
